from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd

inpath = '../inputs/sales_multisheet_excel.xlsx'

# Load in the workbook
wb = load_workbook(inpath)

# Get sheet names
#print(wb.get_sheet_names())
print("\nSheet Names:", wb.sheetnames)

for sheet in wb:           # go over all sheets
    print(sheet.title)


print("\n==== Get a sheet by name or Index ====")
#sheet1 = wb.get_sheet_by_name('Sheet1')
sheet3 = wb['Sheet3']
print("Sheet Type is:", type(sheet3))
print("Sheet Title is:", sheet3.title)
dimensions1 = sheet3.dimensions
print("Sheet 1 dimension is:", dimensions1)
print("sheet.max_row", sheet3.max_row)
print("sheet.max_column", sheet3.max_column)

print("==== Getting Cells from the Sheets ====")
print(sheet3['A8'].value)
print(sheet3['B8'].value)
print(sheet3['C8'].value)
print(sheet3['D8'].value)
print(sheet3['A9'].value)
print(sheet3['B9'].value)
print(sheet3['C9'].value)
print(sheet3['D9'].value)

print("==== Value from rows ====")
rows = [[cell.value for cell in row] for row in sheet3.rows]
num_cols = len(rows[0])
print("Sheet Column Length:", num_cols)
print(rows[0])
print(rows[1])
print(rows[2])
print(rows[3])
print(rows[4])
print(rows[5])
print(rows[6])
print(rows[7])
print("==============")
print(rows[7][0])
print(rows[7][1])
print(rows[7][2])
print(rows[8][0])
print(rows[8][1])
print(rows[8][2])


#print("\n==== # Convert Sheet to DataFrame ====")
#data = sheet.values
# Convert Sheet to DataFrame
#df = pd.DataFrame(sheet3.values)
#df = pd.DataFrame(data, index=13, columns=['Name','Total'])
#print(df)


#print("==== Converting a worksheet to a Dataframe ====")
#data = sheet3.values
#cols = next(data)[1:]
#data = list(data)
#idx = [r[0] for r in data]
#data = (islice(r, 1, None) for r in data)
#df = pd.DataFrame(data, index=idx, columns=cols)
#print(df)
