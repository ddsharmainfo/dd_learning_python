from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

inpath = '../inputs/sales_multisheet_excel.xlsx'

# Load in the workbook
wb = load_workbook(inpath)
print(type(wb))

# Get sheet names
#print(wb.get_sheet_names())
print("\nSheet Names:", wb.sheetnames)

for sheet in wb:           # go over all sheets
    print(sheet.title)


print("\n==== Get a sheet by name or Index ====")
#sheet1 = wb.get_sheet_by_name('Sheet1')
sheet1 = wb['Sheet3']
print("Sheet Type is:", type(sheet1))
print("Sheet Title is:", sheet1.title)
dimensions1 = sheet1.dimensions
print("Sheet 1 dimension is:", dimensions1)
print("sheet.max_row", sheet1.max_row)
print("sheet.max_column", sheet1.max_column)

print("==== Getting Cells from the Sheets ====")
print(sheet1['A1'].value)
print(sheet1['A2'].value)
print(sheet1['B1'].value)
print(sheet1['B2'].value)

temp = sheet1['B1']
temp.value
print('Row ' + str(temp.row) + ', Column ' + temp.column + ' is ' + temp.value)
print('Cell ' + temp.coordinate + ' is ' + temp.value)

print("sheet.cell(row=1, column=2) =", sheet1.cell(row=1, column=2).value)

for i in range(1, 8, 2):
    print(i, sheet1.cell(row=i, column=2).value)

print("==== Converting Between Column Letters and Numbers ====")
#print(get_column_letter(2))


print("\nGetting Rows and Columns from the Sheets")
print(tuple(sheet1['A1':'C3']))

for rowOfCellObjects in sheet1['A1':'B8']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')


for i in range(1, sheet1.max_row):
    for j in range(1, sheet1.max_column):
        print(i, sheet1.cell(row=i, column=j).value)

rows = [[cell.value for cell in row] for row in sheet1.rows]
num_cols = len(rows[0])
print("Sheet Column Length:", num_cols)
print(rows[0])
print(rows[1])
print(rows[2])
print(rows[3])
print(rows[4])
print(rows[5])

# Print the sheet title
sheet1.title

# Get currently active sheet
anotherSheet = wb.active

# Check `anotherSheet`
anotherSheet

# Retrieve the value of a certain cell
sheet1['A1'].value

# Select element 'B2' of your sheet
c = sheet1['B2']

# Retrieve the row number of your element
c.row




print("\n==== Sheet 3 Data ====")
sheet3 = wb.worksheets[2]
dimensions3 = sheet3.dimensions
print(dimensions3)








